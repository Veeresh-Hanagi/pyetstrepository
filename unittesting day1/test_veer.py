import time
import unittest

import openpyxl
from selenium import webdriver
from faker import Faker
fake=Faker()
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from openpyxl import workbook
file="C:\\Users\\Veeresh\\Documents\\repo\\Book.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]
for i in range(1,20):
    sheet.cell(i,1).value=fake.name()
    workbook.save(file)
for i in range(1,20):
    sheet.cell(i,2).value=fake.email()
    workbook.save(file)
for i in range(1,20):
    sheet.cell(i,3).value=fake.basic_phone_number()
    workbook.save(file)
for i in range(1,20):
    sheet.cell(i,4).value=fake.address()
    workbook.save(file)

class features(unittest.TestCase):
    def setUp(self):
        serv_obj=Service("C:\webdrivers\chromedriver-win64\chromedriver-win64\chromedriver.exe")
        self.driver=webdriver.Chrome(service=serv_obj)
        self.driver.get("https://testautomationpractice.blogspot.com/")
        self.driver.maximize_window()
        time.sleep(5)
#
    def test_first_name_field_box(self):
        rows=sheet.max_row
        cols = sheet.max_column
        for i in range(1, rows + 1):
                self.driver.find_element(By.XPATH, "//input[@id='name']").send_keys(sheet.cell(i,1).value)
                time.sleep(4)
                self.driver.find_element(By.XPATH, "//input[@id='email']").send_keys(sheet.cell(i,2).value)
                time.sleep(4)
                self.driver.find_element(By.XPATH, "//input[@id='phone']").send_keys(sheet.cell(i,3).value)
                time.sleep(4)
                self.driver.find_element(By.XPATH, "//textarea[@id='textarea']").send_keys(sheet.cell(i,4).value)
                time.sleep(4)
                self.driver.find_element(By.XPATH, "//input[@id='name']").clear()
                self.driver.find_element(By.XPATH, "//input[@id='email']").clear()
                self.driver.find_element(By.XPATH, "//input[@id='phone']").clear()
                self.driver.find_element(By.XPATH, "//textarea[@id='textarea']").clear()


    def tearDown(self):
        self.driver.close()


